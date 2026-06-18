"""Sinh file Excel mẫu cho Kho Template QA từ data/qa-templates.json.
Chạy: python scripts/gen-templates.py
Output: public/templates/<slug>.xlsx (chỉ cho template dạng bảng).
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data", "qa-templates.json")
OUT = os.path.join(ROOT, "public", "templates")

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="6D28D9")  # brand-700
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, color="1E1B4B", size=13)
CELL_FONT = Font(name=FONT, size=11)
NOTE_FONT = Font(name=FONT, italic=True, color="64748B", size=10)
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Dropdown cho từng tên cột
DROPDOWNS = {
    "Kết quả": ["Chưa test", "Pass", "Fail", "N/A"],
    "Kết quả thực tế": ["Chưa test", "Pass", "Fail", "N/A"],
    "Trạng thái": ["Chưa chạy", "Pass", "Fail", "Blocked"],
    "Trạng thái lỗi": ["Mới", "Đang sửa", "Đã sửa", "Cần retest", "Đóng", "Mở lại"],
    "Mức độ": ["Nghiêm trọng", "Cao", "Trung bình", "Thấp"],
    "Ưu tiên": ["Cao", "Trung bình", "Thấp"],
    "Loại": ["Happy path", "Negative", "Boundary", "Security", "Performance"],
    "Loại rule": ["Tính toán", "Ngưỡng/điều kiện", "Validate", "Phân quyền", "Trạng thái dữ liệu"],
    "Nguồn": ["Code", "Config", "DB", "Feature flag"],
    "Trạng thái xác minh": ["Đã xác nhận", "Nghi vấn", "Nghi là bug", "Cần hỏi BA"],
    "Đánh giá": ["Đạt", "Không đạt", "Cần sửa", "N/A"],
}

WIDE_COLS = {
    "Các bước", "Mô tả ngắn", "Kết quả mong đợi", "Hạng mục kiểm tra",
    "Môi trường / Điểm kiểm tra", "Tiêu đề", "Mô tả requirement",
    "Tên test case", "Tiêu chí nghiệm thu", "Response mong đợi",
    "Payload / Params", "Endpoint", "Header / Auth",
    "Quy tắc nghiệp vụ", "Vị trí trong code", "Câu hỏi cho BA",
    "Hạng mục cần soát",
}

EXTRA_ROWS = 40  # số dòng trống chừa sẵn để điền thêm


def build(template):
    c = template["content"]
    if c["kind"] != "table":
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    cols = c["columns"]
    ncol = len(cols)
    r = 1

    # Tiêu đề / ngữ cảnh
    if c.get("context"):
        ws.cell(row=r, column=1, value=c["context"]).font = TITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        r += 2

    header_row = r
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=r, column=ci, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    r += 1

    first_data_row = r
    for row in c["rows"]:
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
        r += 1

    # Dòng trống chừa sẵn (vẫn kẻ viền + dropdown áp xuống)
    for _ in range(EXTRA_ROWS):
        for ci in range(1, ncol + 1):
            cell = ws.cell(row=r, column=ci)
            cell.font = CELL_FONT
            cell.border = BORDER
        r += 1
    last_row = r - 1

    # Dropdown chọn cho các cột phù hợp
    for ci, col in enumerate(cols, start=1):
        if col in DROPDOWNS:
            letter = get_column_letter(ci)
            opts = ",".join(DROPDOWNS[col])
            dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=True)
            dv.add(f"{letter}{first_data_row}:{letter}{last_row}")
            ws.add_data_validation(dv)

    # Ghi chú dưới bảng
    if c.get("note"):
        cell = ws.cell(row=last_row + 2, column=1, value=c["note"])
        cell.font = NOTE_FONT
        ws.merge_cells(start_row=last_row + 2, start_column=1, end_row=last_row + 2, end_column=ncol)

    # Độ rộng cột
    for ci, col in enumerate(cols, start=1):
        letter = get_column_letter(ci)
        if col in ("ID", "#", "Test Case ID", "Bug ID", "Req ID", "TC ID", "Method"):
            w = 12
        elif col in ("Kết quả", "Kết quả thực tế", "Trạng thái", "Trạng thái lỗi", "Mức độ", "Ưu tiên", "Loại", "Status mong đợi"):
            w = 14
        elif col in WIDE_COLS:
            w = 38
        else:
            w = 20
        ws.column_dimensions[letter].width = w

    # Freeze header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    return wb


def main():
    os.makedirs(OUT, exist_ok=True)
    groups = json.load(open(DATA, encoding="utf-8"))
    made = []
    for g in groups:
        for t in g["templates"]:
            wb = build(t)
            if wb is None:
                continue
            path = os.path.join(OUT, f"{t['slug']}.xlsx")
            wb.save(path)
            made.append(os.path.basename(path))
    print("Created", len(made), "files:")
    for m in made:
        print("  -", m)


if __name__ == "__main__":
    main()
